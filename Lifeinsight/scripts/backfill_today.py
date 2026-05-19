"""5/14（当日）のフォーム回答行をAMeDAS APIデータから補完"""
import json
import datetime
import openpyxl
import sys
sys.path.insert(0, 'scripts')
from backfill import (calc_moon_age, is_holiday, round_to)

XLSX = 'data/Lifeinsight.xlsx'

# AMeDAS 5/14 データを統合
hourly = {}  # hour -> {pressure, temp, humidity}
for h in ['00', '03', '06', '09', '12', '15']:
    with open(f'data/jma/amedas_20260514_{h}.json') as f:
        data = json.load(f)
    for ts, vals in data.items():
        # ts: YYYYMMDDHHmmss, 00分のみ採用
        if ts[10:12] != '00':
            continue
        hh = int(ts[8:10])
        # 0時は前日の24時として扱う
        hour = hh if hh > 0 else 24
        # 5/14のキーのみ
        if ts[:8] != '20260514':
            continue
        hourly[hour] = {
            'pressure': vals.get('pressure', [None])[0],
            'temp': vals.get('temp', [None])[0],
            'humidity': vals.get('humidity', [None])[0],
        }

print('5/14 取得時刻:', sorted(hourly.keys()))

# 5/13 のデータを既存スクリプトから取得（寒暖差計算用）
import sys
sys.path.insert(0, 'scripts')
from backfill import parse_jma_html
data_513 = parse_jma_html('data/jma/2026-05-13.html')
print('5/13 取得時刻:', sorted(data_513.keys()))

# 5/14 12:58 のフォーム入力行を補完
target = datetime.datetime(2026, 5, 14, 12, 58, 48)

# 12:00 と 13:00 のうち、12:58に近い方を採用 → 13:00
def nearest_hour(hourly_dict, hh):
    if hh in hourly_dict and hourly_dict[hh].get('pressure') is not None:
        return hourly_dict[hh]
    for offset in range(1, 6):
        for h in [hh - offset, hh + offset]:
            if h in hourly_dict and hourly_dict[h].get('pressure') is not None:
                return hourly_dict[h]
    return None

v13 = nearest_hour(hourly, 13)
print('5/14 13:00 のデータ:', v13)

# 気圧トレンド: 13:00 - 10:00
v10 = nearest_hour(hourly, 10)
trend = round_to(v13['pressure'] - v10['pressure'], 1) if v10 else None

# 24h変動幅: 5/13 14時〜5/14 13時 (過去24時間)
pressures = []
for h in range(14, 25):
    p = data_513.get(h, {}).get('pressure')
    if p is not None:
        pressures.append(p)
for h in range(1, 14):
    p = hourly.get(h, {}).get('pressure')
    if p is not None:
        pressures.append(p)
p24 = round_to(max(pressures) - min(pressures), 1) if pressures else None

# 寒暖差: 5/14 1〜13時の最高 - 5/13 全時間の最高
today_max = max((v.get('temp') for h, v in hourly.items() if h <= 13 and v.get('temp') is not None), default=None)
yesterday_max = max((v.get('temp') for h, v in data_513.items() if v.get('temp') is not None), default=None)
tdiff = round_to(today_max - yesterday_max, 1) if (today_max is not None and yesterday_max is not None) else None

print(f'計算結果: P={v13["pressure"]}, T={v13["temp"]}, H={v13["humidity"]}, trend={trend}, 24h={p24}, tdiff={tdiff}')

wb = openpyxl.load_workbook(XLSX)
ws = wb['フォームの回答 1']
# 5/14のフォーム行を検索
for row in range(2, ws.max_row + 1):
    ts = ws.cell(row=row, column=1).value
    if not isinstance(ts, datetime.datetime):
        continue
    if ts.date() == datetime.date(2026, 5, 14):
        ws.cell(row=row, column=7).value = round_to(v13['pressure'], 1)
        ws.cell(row=row, column=8).value = trend
        ws.cell(row=row, column=9).value = p24
        ws.cell(row=row, column=10).value = round_to(v13['temp'], 1)
        ws.cell(row=row, column=11).value = tdiff
        ws.cell(row=row, column=12).value = round(v13['humidity']) if v13['humidity'] is not None else None
        ws.cell(row=row, column=13).value = calc_moon_age(ts)
        ws.cell(row=row, column=14).value = is_holiday(ts.date())
        print(f'Row {row} 補完完了')
        break

wb.save(XLSX)
print('保存完了')
