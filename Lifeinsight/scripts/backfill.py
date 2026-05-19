"""
気象庁の過去気象データ（時間別）から欠損気象データを補完してxlsxを更新する。
"""
import os
import math
import datetime
import re
from collections import defaultdict
from bs4 import BeautifulSoup
import openpyxl

JMA_DIR = 'data/jma'
XLSX_PATH = 'data/Lifeinsight.xlsx'

# 日本の祝日（2026年4月〜5月）
HOLIDAYS_2026 = {
    datetime.date(2026, 4, 29),  # 昭和の日
    datetime.date(2026, 5, 3),   # 憲法記念日
    datetime.date(2026, 5, 4),   # みどりの日
    datetime.date(2026, 5, 5),   # こどもの日
    datetime.date(2026, 5, 6),   # 振替休日（5/3が日曜）
}

def parse_jma_html(path):
    """JMAページから時間別データ(時刻->{pressure, temp, humidity})を抽出。
       湿度・気圧が欠測の時刻は値をNoneにして含める。"""
    with open(path, 'r', encoding='utf-8') as f:
        html = f.read()
    soup = BeautifulSoup(html, 'html.parser')
    tables = soup.find_all('table')
    target = None
    for t in tables:
        if 'data2_s' in t.get('class', []) and len(t.find_all('tr')) > 20:
            target = t
            break
    if target is None:
        return {}
    result = {}
    for r in target.find_all('tr')[2:]:  # ヘッダー2行スキップ
        cells = [c.get_text(strip=True) for c in r.find_all(['th', 'td'])]
        if len(cells) < 8:
            continue
        try:
            hour = int(cells[0])
        except ValueError:
            continue
        def parse_num(s):
            try:
                return float(s)
            except (ValueError, TypeError):
                return None
        result[hour] = {
            'pressure': parse_num(cells[1]),  # 現地気圧
            'temp': parse_num(cells[4]),
            'humidity': parse_num(cells[7]),
        }
    return result

def load_all_data(start_date, end_date):
    """指定期間のデータを {date(date型): {hour: {...}}} で返す"""
    data = {}
    d = start_date
    while d <= end_date:
        path = os.path.join(JMA_DIR, d.strftime('%Y-%m-%d') + '.html')
        if os.path.exists(path):
            data[d] = parse_jma_html(path)
        d += datetime.timedelta(days=1)
    return data

def round_to(v, decimals):
    if v is None:
        return None
    factor = 10 ** decimals
    return round(v * factor) / factor

def calc_moon_age(dt):
    """GASと同じ簡易月齢計算"""
    known_new_moon = datetime.datetime(2000, 1, 6, 18, 14, 0)
    synodic = 29.53058867
    diff_days = (dt - known_new_moon).total_seconds() / 86400.0
    age = diff_days % synodic
    if age < 0:
        age += synodic
    return round_to(age, 1)

def is_holiday(date):
    if date.weekday() >= 5:  # 土日
        return True
    return date in HOLIDAYS_2026

def get_hourly_value(data, dt, key):
    """指定日時の最も近い時刻の値を返す。値がNoneなら近傍時刻にフォールバック。"""
    date = dt.date()
    if date not in data:
        return None
    # JMAの「24時」は翌日0時に相当するが、当日24時として扱う
    # 0時のデータは存在しないので、1〜24時の中で最も近いもの
    target_hour = dt.hour if dt.hour > 0 else 24
    # 同じ日のデータから最も近い時刻
    if target_hour in data[date] and data[date][target_hour].get(key) is not None:
        return data[date][target_hour][key]
    # 近傍時刻にフォールバック
    for offset in range(1, 13):
        for h in [target_hour - offset, target_hour + offset]:
            if h in data[date] and data[date][h].get(key) is not None:
                return data[date][h][key]
    return None

def calc_pressure_trend(data, dt):
    """3時間前との気圧差"""
    cur = get_hourly_value(data, dt, 'pressure')
    if cur is None:
        return None
    prev_dt = dt - datetime.timedelta(hours=3)
    prev = get_hourly_value(data, prev_dt, 'pressure')
    if prev is None:
        return None
    return round_to(cur - prev, 1)

def calc_pressure_24h_range(data, dt, confirmed=False):
    """confirmed=True: 当日全時間帯の気圧最大-最小
       confirmed=False: 過去24時間の気圧最大-最小"""
    pressures = []
    if confirmed:
        date = dt.date()
        if date in data:
            for h in range(1, 25):
                v = data[date].get(h, {}).get('pressure')
                if v is not None:
                    pressures.append(v)
    else:
        # 過去24時間
        for h_offset in range(0, 25):
            t = dt - datetime.timedelta(hours=h_offset)
            v = get_hourly_value(data, t, 'pressure')
            if v is not None:
                pressures.append(v)
    if not pressures:
        return None
    return round_to(max(pressures) - min(pressures), 1)

def calc_temp_diff(data, dt, confirmed=False):
    """当日最高気温 - 前日最高気温"""
    today = dt.date()
    yesterday = today - datetime.timedelta(days=1)
    today_max = None
    yesterday_max = None
    if today in data:
        for h, v in data[today].items():
            t = v.get('temp')
            if t is None:
                continue
            if confirmed or h <= dt.hour:
                today_max = t if today_max is None else max(today_max, t)
    if yesterday in data:
        for h, v in data[yesterday].items():
            t = v.get('temp')
            if t is None:
                continue
            yesterday_max = t if yesterday_max is None else max(yesterday_max, t)
    if today_max is None or yesterday_max is None:
        return None
    return round_to(today_max - yesterday_max, 1)

# ============================================================
# メイン処理
# ============================================================

def main():
    # データ読み込み: 4/9 〜 5/14
    data = load_all_data(datetime.date(2026, 4, 9), datetime.date(2026, 5, 14))
    print(f'Loaded {len(data)} days')
    for d in sorted(data.keys()):
        hours_with_pressure = sum(1 for h in data[d].values() if h.get('pressure') is not None)
        print(f'  {d}: {len(data[d])} hours, {hours_with_pressure} with pressure')

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws_form = wb['フォームの回答 1']
    ws_log = wb['気象ログ']

    # --- フォーム回答シートの欠損補完 ---
    print('\n=== フォーム回答シートの補完 ===')
    form_dates = set()
    for row in range(2, ws_form.max_row + 1):
        ts = ws_form.cell(row=row, column=1).value
        if not isinstance(ts, datetime.datetime):
            continue
        form_dates.add(ts.date())
        # 7列目（瞬間気圧）が空ならば補完
        if ws_form.cell(row=row, column=7).value is not None:
            continue
        # 計算
        pressure = get_hourly_value(data, ts, 'pressure')
        temp = get_hourly_value(data, ts, 'temp')
        humidity = get_hourly_value(data, ts, 'humidity')
        trend = calc_pressure_trend(data, ts)
        p24 = calc_pressure_24h_range(data, ts, confirmed=False)
        tdiff = calc_temp_diff(data, ts, confirmed=False)
        moon = calc_moon_age(ts)
        holiday = is_holiday(ts.date())

        ws_form.cell(row=row, column=7).value = round_to(pressure, 1)
        ws_form.cell(row=row, column=8).value = trend
        ws_form.cell(row=row, column=9).value = p24
        ws_form.cell(row=row, column=10).value = round_to(temp, 1)
        ws_form.cell(row=row, column=11).value = tdiff
        ws_form.cell(row=row, column=12).value = round(humidity) if humidity is not None else None
        ws_form.cell(row=row, column=13).value = moon
        ws_form.cell(row=row, column=14).value = holiday
        print(f'  Row {row} ({ts.strftime("%Y-%m-%d %H:%M")}): P={pressure}, T={temp}, H={humidity}, trend={trend}, 24h={p24}, tdiff={tdiff}, moon={moon}, holiday={holiday}')

    # --- 気象ログシートの補完 ---
    print('\n=== 気象ログシートの補完 ===')
    # 既存日付収集 + 5/4は再計算対象
    existing_log_dates = {}
    for row in range(2, ws_log.max_row + 1):
        d = ws_log.cell(row=row, column=1).value
        if isinstance(d, datetime.datetime):
            existing_log_dates[d.date()] = row
        elif isinstance(d, datetime.date):
            existing_log_dates[d] = row

    # 5/4の不完全行を再計算で上書き
    target_date = datetime.date(2026, 5, 4)
    if target_date in existing_log_dates:
        row = existing_log_dates[target_date]
        rep_dt = datetime.datetime(2026, 5, 4, 12, 0, 0)
        pressure = get_hourly_value(data, rep_dt, 'pressure')
        temp = get_hourly_value(data, rep_dt, 'temp')
        humidity = get_hourly_value(data, rep_dt, 'humidity')
        trend = calc_pressure_trend(data, rep_dt)
        p24 = calc_pressure_24h_range(data, rep_dt, confirmed=True)
        tdiff = calc_temp_diff(data, rep_dt, confirmed=True)
        moon = calc_moon_age(datetime.datetime(2026, 5, 4))
        holiday = is_holiday(target_date)
        ws_log.cell(row=row, column=2).value = round_to(pressure, 1)
        ws_log.cell(row=row, column=3).value = trend
        ws_log.cell(row=row, column=4).value = p24
        ws_log.cell(row=row, column=5).value = round_to(temp, 1)
        ws_log.cell(row=row, column=6).value = tdiff
        ws_log.cell(row=row, column=7).value = round(humidity) if humidity is not None else None
        ws_log.cell(row=row, column=8).value = moon
        ws_log.cell(row=row, column=9).value = holiday
        print(f'  Row {row} ({target_date}): 再計算 P={pressure}, T={temp}, H={humidity}')

    # フォーム入力なし & 気象ログにない日を追記
    d = datetime.date(2026, 4, 10)
    end = datetime.date(2026, 5, 13)
    while d <= end:
        if d in form_dates or d in existing_log_dates:
            d += datetime.timedelta(days=1)
            continue
        rep_dt = datetime.datetime(d.year, d.month, d.day, 12, 0, 0)
        pressure = get_hourly_value(data, rep_dt, 'pressure')
        temp = get_hourly_value(data, rep_dt, 'temp')
        humidity = get_hourly_value(data, rep_dt, 'humidity')
        trend = calc_pressure_trend(data, rep_dt)
        p24 = calc_pressure_24h_range(data, rep_dt, confirmed=True)
        tdiff = calc_temp_diff(data, rep_dt, confirmed=True)
        moon = calc_moon_age(datetime.datetime(d.year, d.month, d.day))
        holiday = is_holiday(d)
        new_row = ws_log.max_row + 1
        ws_log.cell(row=new_row, column=1).value = datetime.datetime(d.year, d.month, d.day)
        ws_log.cell(row=new_row, column=2).value = round_to(pressure, 1)
        ws_log.cell(row=new_row, column=3).value = trend
        ws_log.cell(row=new_row, column=4).value = p24
        ws_log.cell(row=new_row, column=5).value = round_to(temp, 1)
        ws_log.cell(row=new_row, column=6).value = tdiff
        ws_log.cell(row=new_row, column=7).value = round(humidity) if humidity is not None else None
        ws_log.cell(row=new_row, column=8).value = moon
        ws_log.cell(row=new_row, column=9).value = holiday
        print(f'  追加 Row {new_row} ({d}): P={pressure}, T={temp}, H={humidity}')
        d += datetime.timedelta(days=1)

    wb.save(XLSX_PATH)
    print(f'\n保存完了: {XLSX_PATH}')

if __name__ == '__main__':
    main()
